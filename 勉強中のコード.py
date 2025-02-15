#勉強中のコード

import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import re

def extract_runtime(soup):
    """
    様々な形式の再生時間情報（例："1時間20分", "20分", "01:20:00"）に対応して抽出する関数。
    該当する<td>タグやテキストがあれば、合計分数を文字列で返す。
    """
    # パターン1: "1時間20分"または"20分"形式
    time_tag = soup.find(lambda tag: tag.name == "td" and re.search(r'(\d+\s*時間\s*)?\d+\s*分', tag.text))
    if time_tag:
        raw_time = time_tag.text.strip()
        hm_match = re.search(r'(?:(\d+)\s*時間)?\s*(\d+)\s*分', raw_time)
        if hm_match:
            hours = hm_match.group(1)
            minutes = hm_match.group(2)
            total_minutes = (int(hours) * 60 if hours else 0) + int(minutes)
            return str(total_minutes)
    
    # パターン2: "hh:mm(:ss)"形式
    time_tag = soup.find(lambda tag: tag.name == "td" and re.search(r'\d+:\d+(?::\d+)?', tag.text))
    if time_tag:
        raw_time = time_tag.text.strip()
        time_match = re.search(r'(\d+):(\d+)(?::(\d+))?', raw_time)
        if time_match:
            hours, minutes, _ = time_match.groups()
            total_minutes = int(hours) * 60 + int(minutes)
            return str(total_minutes)
    
    return "時間情報なし"

def download_data(url, excel_path):
    # ヘッダーとクッキーの設定
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
    }
    cookie = {'age_check_done': '1'}

    # 固定のURLリストファイルパス
    url_list_path = "" #固定URLリストの入力

    # セッションの初期化
    session = requests.Session()
    session.get(url, headers=headers, cookies=cookie)

    # Excelファイルのロード
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb['記事作成部分']
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの読み込みに失敗しました: {e}")
        return

    row_num = 4  # 書き込み開始行

    try:
        with open(url_list_path, 'r', encoding='UTF-8') as f:
            urls = f.readlines()
    except Exception as e:
        messagebox.showerror("エラー", f"固定URLリストファイルの読み込みに失敗しました: {e}")
        return

    for idx, line in enumerate(reversed(urls), start=1):
        page_url = line.strip()
        print(f"Processing URL ({idx}): {page_url}")

        try:
            # HTMLの取得と解析
            response = session.get(page_url, headers=headers, cookies=cookie)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')

            # データの抽出
            url_tag = soup.find('meta', property="og:url")
            url_title = soup.find(id='title')
            performer = soup.find(id='performer')
            runtime = extract_runtime(soup)

            url_tag_content = url_tag['content'] if url_tag else "URLが見つかりません"
            url_title_text = url_title.text.strip() if url_title else "タイトルが見つかりません"
            performer_text = performer.text.strip() if performer else "出演者情報なし"

            # Excelに書き込み
            sheet.cell(row=row_num, column=3, value=url_title_text)
            sheet.cell(row=row_num, column=4, value=runtime)
            sheet.cell(row=row_num, column=5, value=url_tag_content)
            sheet.cell(row=row_num, column=6, value=performer_text)

            row_num += 1

        except Exception as e:
            print(f"URL {page_url} の処理中にエラーが発生しました: {e}")
            continue

    # Excelファイルを保存
    try:
        wb.save(excel_path)
        messagebox.showinfo("完了", f"Excelファイルを保存しました: {excel_path}")
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの保存に失敗しました: {e}")

def select_file_and_run(root):
    # Excelファイル選択
    excel_path = filedialog.askopenfilename(
        title="Excelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx;*.xls")],
        initialdir=os.path.expanduser("~/Documents")
    )

    if not excel_path:
        messagebox.showwarning("警告", "Excelファイルが選択されていません")
        return

    # 確認メッセージ
    confirm_msg = (
        f"以下のExcelファイルで処理を開始しますか？\n\n"
        f"Excelファイル: {excel_path}\n"
        f"固定URLリスト: C:\\Users\\yuruy\\Downloads\\AV\\FANZA_URL.txt"
    )
    if not messagebox.askyesno("確認", confirm_msg):
        return

    # GUIを閉じる
    root.destroy()

    try:
        url_certification = 'https://www.dmm.co.jp/age_check/=/declared=yes/?rurl=https%3A%2F%2Fwww.dmm.co.jp%2Ftop%2F'
        download_data(url_certification, excel_path)
        sys.exit()
    except Exception as e:
        messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")
        sys.exit()

def main():
    root = tk.Tk()
    root.title("FANZA情報スクレイピング")
    root.attributes("-topmost", True)
    root.geometry("400x300")

    label = tk.Label(root, text="Excelファイルを選択してください", font=("Arial", 12))
    label.pack(pady=20)

    button = tk.Button(root, text="ファイルを選択して実行", command=lambda: select_file_and_run(root), font=("Arial", 12), bg="lightblue")
    button.pack(pady=40)

    root.mainloop()

if __name__ == "__main__":
    main()
