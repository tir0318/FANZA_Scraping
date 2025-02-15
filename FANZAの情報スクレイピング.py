#FANZAの情報スクレイピング　現在

import requests
from bs4 import BeautifulSoup
import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import sys

def download_images(url, excel_path):
    # ヘッダーとクッキーの設定
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"
    }
    cookie = {'age_check_done': '1'}

    # 固定のURLリストファイルパス
    url_list_path = "" #固定URLリストを入力
    # セッションの初期化
    session = requests.Session()
    session.get(url, headers=headers, cookies=cookie)

    # Excelファイルのロード
    try:
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb['記事作成部分']
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの読み込みに失敗しました: {e}")
        return

    row_num = 4  # 開始行
    try:
        with open(url_list_path, 'r', encoding='UTF-8') as f:
            urls = f.readlines()
    except Exception as e:
        messagebox.showerror("エラー", f"固定URLリストファイルの読み込みに失敗しました: {e}")
        return

    for idx, line in enumerate(reversed(urls), start=1):
        page_url = line.strip()
        print(f"Processing URL ({idx}): {page_url}")

        try:
            # HTMLの取得と解析
            response = session.get(page_url, headers=headers, cookies=cookie)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, 'html.parser')

            # データの抽出
            url_tag = soup.find('meta', property="og:url")
            url_title = soup.find(id='title')
            name = soup.find(id='performer')
            time_tag = soup.find_all('td')

            # 要素の存在確認とデータ取得
            url_tag_content = url_tag['content'] if url_tag else "URLが見つかりません"
            url_title_text = url_title.text.strip() if url_title else "タイトルが見つかりません"
            name_text = name.text.strip() if name else "出演者情報なし"
            time_text = time_tag[10].text.strip()[:-1] if len(time_tag) > 10 else "時間情報なし"

            # Excelに書き込み
            sheet.cell(row=row_num, column=3, value=url_title_text)
            sheet.cell(row=row_num, column=4, value=time_text)
            sheet.cell(row=row_num, column=5, value=url_tag_content)
            sheet.cell(row=row_num, column=6, value=name_text)

            row_num += 1

        except Exception as e:
            print(f"URL {page_url} の処理中にエラーが発生しました: {e}")
            continue

    # Excelファイルを保存
    try:
        wb.save(excel_path)
        messagebox.showinfo("完了", f"Excelファイルを保存しました: {excel_path}")
    except Exception as e:
        messagebox.showerror("エラー", f"Excelファイルの保存に失敗しました: {e}")

def select_file_and_run(root):
    # Excelファイル選択
    excel_path = filedialog.askopenfilename(
        title="Excelファイルを選択してください",
        filetypes=[("Excel Files", "*.xlsx;*.xls")],
        initialdir=os.path.expanduser("~/Documents")
    )

    if not excel_path:
        messagebox.showwarning("警告", "Excelファイルが選択されていません")
        return

    # 確認メッセージ
    if not messagebox.askyesno("確認", f"以下のExcelファイルで処理を開始しますか？\n\nExcelファイル: {excel_path}\n固定URLリスト: C:\\Users\\yuruy\\Downloads\\AV\\FANZA_URL.txt"):
        return

    # GUIを閉じる
    root.destroy()

    # ダウンロード処理を実行
    try:
        url_certification = 'https://www.dmm.co.jp/age_check/=/declared=yes/?rurl=https%3A%2F%2Fwww.dmm.co.jp%2Ftop%2F'
        download_images(url_certification, excel_path)
        sys.exit()
    except Exception as e:
        messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{e}")
        sys.exit()

def main():
    # Tkinterウィンドウを設定
    root = tk.Tk()
    root.title("FANZA情報スクレイピング")

    # ウィンドウを常に最前面に設定
    root.attributes("-topmost", True)

    # ウィンドウのサイズを設定
    root.geometry("400x300")

    # 説明ラベル
    label = tk.Label(root, text="Excelファイルを選択してください", font=("Arial", 12))
    label.pack(pady=20)

    # 実行ボタン
    button = tk.Button(root, text="ファイルを選択して実行", command=lambda: select_file_and_run(root), font=("Arial", 12), bg="lightblue")
    button.pack(pady=40)

    # Tkinterメインループ開始
    root.mainloop()

if __name__ == "__main__":
    main()